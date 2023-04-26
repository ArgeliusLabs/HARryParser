import openpyxl
import tldextract

try:
    from .helpers import *
    from .logger import logger
except:
    from helpers import *
    from logger import logger

from dns_helper import SubdomainScanner


class HarParser():

    def __init__(self, har_file, output_file, parent_domain, tds_file="tds.json", check_dns=True):
        self.har_file = har_file
        self.output_file = output_file
        self.parent_domain = parent_domain
        self.tds_file = tds_file
        self.check_dns = check_dns

        self.headers = []
        self.cookies = []
        self.querystring = []
        self.postdata = []
        self.parsed_postdata = []
        self.subdomains = []

    def create_workbook(self):
        columns = ["url", "request_id", "name", "value", "calculated_domain", "calculated_entity", "source"]
        wb = openpyxl.Workbook()

        # Remove the default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        parent_sheet = wb.create_sheet("Everything")
        parent_row = 0
        for sheet_name, data in zip(["Headers", "Cookies", "QueryString", "PostData", "ParsedPostData"],
                                    (
                                            self.headers, self.cookies, self.querystring, self.postdata,
                                            self.parsed_postdata)):

            sheet = wb.create_sheet(sheet_name)
            for i in range(len(columns)):
                sheet.cell(row=1, column=i + 1).value = columns[i]
                if parent_row == 0:
                    parent_sheet.cell(row=1, column=i + 1).value = columns[i]

            for row in range(len(data)):
                parent_row += 1
                for col in range(len(columns)):
                    try:
                        sheet.cell(row=row + 2, column=col + 1).value = remove_illegal_chars(data[row][col])
                        parent_sheet.cell(row=parent_row + 1, column=col + 1).value = remove_illegal_chars(
                            data[row][col])
                    except:
                        print("ERROR ON ROW", data[row][col])
                        print("ERROR ON ROW FULL", data[row])
                        traceback.print_exc()

        if self.check_dns:
            subs_sheet = wb.create_sheet("DNS")
            scanner = SubdomainScanner(self.subdomains)
            scanner.scan_subdomains()
            data = scanner.results
            rows = []

            # Iterate over each subdomain and record type in the data dictionary
            for subdomain, record_types in data.items():
                for record_type, record_data in record_types.items():
                    # Handle empty record types
                    if not record_data:
                        rows.append({'Subdomain': subdomain, 'Record Type': record_type})
                        continue
                    for ip_address, ip_data in record_data.items():
                        # Handle empty IP addresses
                        if not ip_data:
                            rows.append({'Subdomain': subdomain, 'Record Type': record_type, 'IP Address': ip_address})
                            continue
                        for key, value in ip_data.items():
                            # Handle empty key-value pairs
                            if not value:
                                rows.append(
                                    {'Subdomain': subdomain, 'Record Type': record_type, 'IP Address': ip_address,
                                     'Key': key})
                                continue
                            # Handle multiple redirects
                            if key == 'redirects':
                                for redirect in value:
                                    rows.append(
                                        {'Subdomain': subdomain, 'Record Type': record_type, 'IP Address': ip_address,
                                         'Source IP': redirect[0], 'Target URL': redirect[1]})
                            # Handle multiple A records
                            elif key == 'a_records':
                                # Get resolved_ips for Source IP field
                                resolved_ips = ', '.join(ip_data.get('resolved_ips', []))
                                for item in value:
                                    rows.append(
                                        {'Subdomain': subdomain, 'Record Type': record_type, 'IP Address': ip_address,
                                         'Key': key.capitalize(), 'Value': item, 'Source IP': resolved_ips})
                            # Handle resolved IPs for NS records
                            elif key == 'resolved_ips':
                                # Skip resolved_ips rows
                                continue
            # Write the header row
            header_row = ['Subdomain', 'Record Type', 'IP Address', 'Key', 'Value', 'Source IP', 'Target URL']
            header = [cell for cell in header_row]
            subs_sheet.append(header)
            # Write the data rows
            for row, data_row in enumerate(rows, start=2):
                subs_sheet.cell(row=row, column=1, value=data_row.get('Subdomain', ''))
                subs_sheet.cell(row=row, column=2, value=data_row.get('Record Type', ''))
                subs_sheet.cell(row=row, column=3, value=data_row.get('IP Address', ''))
                subs_sheet.cell(row=row, column=4, value=data_row.get('Key', ''))
                subs_sheet.cell(row=row, column=5, value=data_row.get('Value', ''))
                subs_sheet.cell(row=row, column=6, value=data_row.get('Source IP', ''))
                subs_sheet.cell(row=row, column=7, value=data_row.get('Target URL', ''))
        wb.save(self.output_file)

    def extract_entries(self):
        with open(self.har_file, "r", encoding="utf-8") as f:
            har = json.load(f)

        entries = har.get("log")["entries"]
        tds_domains = fetch_tds(self.tds_file)

        request_id = 0
        for entry in entries:
            url = entry["request"]["url"]
            extracted = tldextract.extract(url)
            domain = extracted.domain + '.' + extracted.suffix
            entity = (lambda s, d, sub: d.get(s,
                                              "No result detected") if s in d.keys() or sub not in s else self.parent_domain)(
                domain, tds_domains, self.parent_domain)
            if entity == self.parent_domain or entity == "No result detected":
                if extracted.subdomain:
                    subdomain = extracted.subdomain + '.' + extracted.domain + '.' + extracted.suffix
                    if not subdomain in self.subdomains:
                        self.subdomains.append(subdomain)
            request_id += 1
            for request_component, data in zip(["headers", "cookies", "queryString"],
                                               (self.headers, self.cookies, self.querystring)):

                for component_entry in entry["request"][request_component]:

                    data.append([url, request_id, component_entry["name"], component_entry["value"], domain, entity,
                                 request_component])
                    analyzed_value = analyze_string(component_entry["value"])
                    if analyzed_value:
                        data.append(
                            [url, request_id, f'!_analyzed_{component_entry["name"]}', analyzed_value, domain, entity,
                             request_component])

            if entry["request"].get("postData"):
                mimeType = entry["request"].get("postData").get("mimeType")
                text = entry["request"].get("postData").get("text")
                self.postdata.append([url, request_id, mimeType, text, domain, entity, "postData"])
                analyzed_value = analyze_string(text)
                if analyzed_value:
                    self.postdata.append([url, request_id, f'!_analyzed_{mimeType}', analyzed_value, domain, entity,
                                          "postData"])

                if "multipart/form-data" in mimeType:
                    parsed_multipart = parse_multipart_form(mimeType, text)
                    if parsed_multipart:
                        for k, v in parsed_multipart.items():
                            self.parsed_postdata.append([url, request_id, k, v, domain, entity, "parsed_postData"])
                            analyzed_value = analyze_string(v)
                            if analyzed_value:
                                self.parsed_postdata.append(
                                    [url, request_id, f'!_analyzed_{k}', analyzed_value, domain, entity,
                                     "parsed_postData"])
                elif text.startswith("{") or text.startswith("[") or type(text) == dict or type(text) == list:
                    try:
                        _dict = text
                        if type(text) == str:
                            _dict = json.loads(text)
                        if type(_dict) == list:
                            # if we end up with a list try to parse the first entry, otherwise it'll bail out
                            _dict = _dict[0]
                        for _k, _v in flatten_dict(_dict).items():

                            self.parsed_postdata.append([url, request_id, _k, _v, domain, entity, "parsed_postData"])
                            analyzed_value = analyze_string(_v)
                            if analyzed_value:
                                self.parsed_postdata.append(
                                    [url, request_id, f'!_analyzed_{_k}', analyzed_value, domain, entity,
                                     "parsed_postData"])
                    except:
                        pass
                elif "text/plain" in mimeType:
                    parsed_code_arguments = parse_post_body_code_arguments(entry["request"].get("postData"))
                    if parsed_code_arguments:
                        for k, v in parsed_code_arguments.items():
                            self.parsed_postdata.append([url, request_id, k, v, domain, entity, "parsed_postData"])
                            analyzed_value = analyze_string(v)
                            if analyzed_value:
                                self.parsed_postdata.append(
                                    [url, request_id, f'!_analyzed_{k}', analyzed_value, domain, entity,
                                     "parsed_postData"])
                else:
                    pass
        self.create_workbook()


import os
import argparse


def main():
    parser = argparse.ArgumentParser(description="HAR Parser CLI")
    parser.add_argument("har_file", type=str, help="HAR file to parse")
    parser.add_argument("-p", "--parent_domain", type=str, required=True, help="Parent domain")
    parser.add_argument("-t", "--tds_file", type=str, default="harryparser/tds.json",
                        help="TDS JSON file (default: tds.json if present)")
    parser.add_argument("-c", "--check_dns", action="store_true", default=True, help="Enable DNS checks")
    parser.add_argument("-o", "--output_dir", type=str, default="output", help="Output directory path")

    args = parser.parse_args()
    # Check if output directory exists and create it if it doesn't
    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)
    output_file = os.path.join(args.output_dir,
                               f"harryparser_{args.parent_domain}_{datetime.now().strftime('%Y-%m-%d_%H.%M.%S')}.xlsx")
    h = HarParser(args.har_file, output_file, args.parent_domain, args.tds_file, check_dns=args.check_dns)
    h.extract_entries()


if __name__ == "__main__":
    main()
